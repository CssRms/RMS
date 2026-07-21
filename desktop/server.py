"""
ZKTeco Local ADMS Server
Receives attendance punches from the device over LAN, saves to SQLite,
broadcasts live updates via SocketIO, and can sync to Railway cloud.

Run directly:  python server.py
Or launch via: python app.py  (system tray mode)
"""
import os
import sys
import datetime
import urllib.parse
import urllib.request
import io
import socket as _socket

from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO
from sqlalchemy import select, update

# ── Template path (changes when frozen by PyInstaller) ────────────────────────
if getattr(sys, 'frozen', False):
    _tmpl = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=_tmpl)
else:
    app = Flask(__name__)

socketio = SocketIO(app, cors_allowed_origins="*")

# ── Config (loaded from settings.json, overridable by app.py) ─────────────────
from config_util import load as _load_cfg, save as _save_cfg, DATA_DIR
from database import engine, employees, attendance

_cfg = _load_cfg()
RAILWAY_URL   = _cfg.get('railway_url', 'https://cssgrouprms.com')
DEVICE_SERIAL = 'LOCAL_SYNC'

# ── Status maps ───────────────────────────────────────────────────────────────
STATUS_MAP    = {0: "Check In", 1: "Check Out", 2: "Break Out",
                 3: "Break In", 4: "OT In",    5: "OT Out"}
STATUS_TO_INT = {v: k for k, v in STATUS_MAP.items()}

# ── Device state ──────────────────────────────────────────────────────────────
device = {"serial": None, "firmware": None, "ip": None, "last_seen": None, "online": False}

def mark_device_seen(sn=None, firmware=None, ip=None):
    device["last_seen"] = datetime.datetime.now()
    device["online"]    = True
    if sn:       device["serial"]   = sn
    if firmware: device["firmware"] = firmware
    if ip:       device["ip"]       = ip

def secs_since_device():
    if not device["last_seen"]:
        return None
    return (datetime.datetime.now() - device["last_seen"]).total_seconds()

# ── REST: general ──────────────────────────────────────────────────────────────

@app.route("/")
def home():
    return render_template("dashboard.html")

@app.route("/api/device")
def api_device():
    info = dict(device)
    if info["last_seen"]:
        info["last_seen"] = info["last_seen"].isoformat()
    ago = secs_since_device()
    info["online"] = ago is not None and ago < 90
    return jsonify(info)

@app.route("/api/config", methods=["GET"])
def api_get_config():
    cfg = _load_cfg()
    return jsonify({k: cfg[k] for k in ("port", "railway_url", "auto_open")})

@app.route("/api/config", methods=["POST"])
def api_save_config():
    global RAILWAY_URL
    data = request.json or {}
    cfg  = _load_cfg()
    if "railway_url" in data:
        cfg["railway_url"] = data["railway_url"].strip()
        RAILWAY_URL = cfg["railway_url"]
    if "port" in data:
        try:
            cfg["port"] = int(data["port"])
        except ValueError:
            return jsonify({"ok": False, "error": "Invalid port"}), 400
    if "auto_open" in data:
        cfg["auto_open"] = bool(data["auto_open"])
    _save_cfg(cfg)
    return jsonify({"ok": True, "message": "Saved. Restart the app to apply a port change."})

@app.route("/api/employees")
def api_employees():
    with engine.connect() as conn:
        rows = conn.execute(select(employees)).fetchall()
        return jsonify([dict(r._mapping) for r in rows])

@app.route("/api/employees", methods=["POST"])
def api_add_employee():
    data = request.json or {}
    if not data.get("staff_id") or not data.get("name"):
        return jsonify({"error": "staff_id and name are required"}), 400
    with engine.connect() as conn:
        try:
            conn.execute(employees.insert().values(
                staff_id=str(data["staff_id"]).strip(),
                name=data["name"].strip(),
                department=data.get("department", ""),
                position=data.get("position", ""),
            ))
            conn.commit()
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 400

@app.route("/api/employees/<staff_id>", methods=["DELETE"])
def api_delete_employee(staff_id):
    with engine.connect() as conn:
        conn.execute(employees.delete().where(employees.c.staff_id == staff_id))
        conn.commit()
    return jsonify({"ok": True})

@app.route("/api/attendance/today")
def api_today():
    today = datetime.date.today()
    start = datetime.datetime(today.year, today.month, today.day)
    end   = start + datetime.timedelta(days=1)
    with engine.connect() as conn:
        rows = conn.execute(
            select(attendance)
            .where(attendance.c.time >= start)
            .where(attendance.c.time < end)
            .order_by(attendance.c.time.desc())
        ).fetchall()
        result = []
        for r in rows:
            d = dict(r._mapping)
            d["time"] = d["time"].isoformat() if d["time"] else None
            result.append(d)
        return jsonify(result)

# ── Diagnostics ────────────────────────────────────────────────────────────────

@app.route("/api/diagnose")
def api_diagnose():
    checks = []

    # 1. Device heartbeat
    ago = secs_since_device()
    if ago is None:
        checks.append({"name": "ZKTeco Device", "status": "error",
                        "message": "No connection yet. Set device Server Address to this laptop's LAN IP and reboot the device."})
    elif ago < 90:
        checks.append({"name": "ZKTeco Device", "status": "ok",
                        "message": f"Online — last heartbeat {int(ago)}s ago  (SN: {device['serial'] or '?'}  IP: {device['ip'] or '?'})"})
    else:
        checks.append({"name": "ZKTeco Device", "status": "warn",
                        "message": f"No heartbeat for {int(ago)}s — device may be offline. Check power and Wi-Fi."})

    # 2. LAN IP (for device config)
    try:
        lan_ip = _socket.gethostbyname(_socket.gethostname())
        port   = _load_cfg().get("port", 80)
        checks.append({"name": "Laptop Network", "status": "ok",
                        "message": f"LAN IP: {lan_ip}  — set this as the device Server Address with Port {port}"})
    except Exception as e:
        checks.append({"name": "Laptop Network", "status": "warn", "message": str(e)})

    # 3. Employees + attendance
    try:
        with engine.connect() as conn:
            emp_count = len(conn.execute(select(employees)).fetchall())
            today     = datetime.date.today()
            start     = datetime.datetime(today.year, today.month, today.day)
            att_today = len(conn.execute(select(attendance).where(attendance.c.time >= start)).fetchall())
            pending   = len(conn.execute(select(attendance).where(attendance.c.synced == 0)).fetchall())
        if emp_count == 0:
            checks.append({"name": "Employee Database", "status": "warn",
                            "message": "No employees registered. Use the dashboard to add employees before starting attendance."})
        else:
            checks.append({"name": "Employee Database", "status": "ok",
                            "message": f"{emp_count} employee(s) registered  ·  {att_today} scan(s) today  ·  {pending} pending cloud sync"})
    except Exception as e:
        checks.append({"name": "Employee Database", "status": "error", "message": f"DB error: {e}"})

    # 4. Railway cloud
    try:
        req = urllib.request.Request(f"{RAILWAY_URL}/iclock/getrequest?SN=DIAGNOSE", method="GET")
        with urllib.request.urlopen(req, timeout=8) as resp:
            if resp.status == 200:
                checks.append({"name": "Railway Cloud", "status": "ok",
                                "message": f"Reachable: {RAILWAY_URL}"})
            else:
                checks.append({"name": "Railway Cloud", "status": "warn",
                                "message": f"Responded with HTTP {resp.status}"})
    except Exception as e:
        checks.append({"name": "Railway Cloud", "status": "warn",
                        "message": f"Not reachable ({type(e).__name__}). You can still capture locally and sync later."})

    # 5. Excel library
    try:
        import openpyxl  # noqa
        checks.append({"name": "Excel Export", "status": "ok", "message": "openpyxl available — Excel export ready"})
    except ImportError:
        checks.append({"name": "Excel Export", "status": "warn",
                        "message": "openpyxl not installed. Run: pip install openpyxl"})

    overall = "ok" if all(c["status"] == "ok" for c in checks) else \
              "error" if any(c["status"] == "error" for c in checks) else "warn"
    return jsonify({"overall": overall, "checks": checks, "timestamp": datetime.datetime.now().isoformat()})

# ── Excel export ───────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
def api_export():
    date_str = request.args.get("date", datetime.date.today().isoformat())
    try:
        target = datetime.date.fromisoformat(date_str)
    except ValueError:
        return jsonify({"error": "Invalid date"}), 400

    start = datetime.datetime(target.year, target.month, target.day)
    end   = start + datetime.timedelta(days=1)
    with engine.connect() as conn:
        rows = conn.execute(
            select(attendance).where(attendance.c.time >= start).where(attendance.c.time < end)
            .order_by(attendance.c.time)
        ).fetchall()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance {date_str}"

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Attendance Report — {target.strftime('%A, %d %B %Y')}"
        ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="0F4C2A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        hdrs = ["Staff ID", "Name", "Department", "Position", "Time", "Status", "Synced"]
        hfill = PatternFill("solid", fgColor="1A7A3E")
        thin  = Side(style="thin", color="CCCCCC")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = hfill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = bdr

        afill = PatternFill("solid", fgColor="F0FAF4")
        for i, r in enumerate(rows):
            d   = dict(r._mapping)
            t   = d["time"]
            row = i + 3
            vals = [
                d["staff_id"], d["name"], d.get("department",""), d.get("position",""),
                t.strftime("%H:%M:%S") if t else "", d.get("status",""),
                "Yes" if d.get("synced") else "Pending"
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="left")
                if i % 2 == 0:
                    cell.fill = afill
                if c == 6 and v and "Out" in v:
                    cell.font = Font(color="C45000")
                if c == 7:
                    cell.font = Font(color="0F4C2A" if v == "Yes" else "BB8800")

        sr = len(rows) + 3
        ws.merge_cells(f"A{sr}:G{sr}")
        ws[f"A{sr}"]           = f"Total: {len(rows)} records"
        ws[f"A{sr}"].font      = Font(bold=True)
        ws[f"A{sr}"].fill      = PatternFill("solid", fgColor="E8F5E9")
        ws[f"A{sr}"].alignment = Alignment(horizontal="right")

        for col, w in zip("ABCDEFG", [14, 22, 16, 16, 10, 12, 10]):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=f"attendance_{date_str}.xlsx")
    except ImportError:
        # Fallback: CSV
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Staff ID","Name","Department","Position","Time","Status","Synced"])
        for r in rows:
            d = dict(r._mapping); t = d["time"]
            w.writerow([d["staff_id"],d["name"],d.get("department",""),d.get("position",""),
                        t.strftime("%H:%M:%S") if t else "",d.get("status",""),
                        "Yes" if d.get("synced") else "Pending"])
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            mimetype="text/csv", as_attachment=True, download_name=f"attendance_{date_str}.csv")

# ── Railway sync ───────────────────────────────────────────────────────────────

@app.route("/api/sync", methods=["POST"])
def api_sync():
    with engine.connect() as conn:
        unsynced = conn.execute(
            select(attendance).where(attendance.c.synced == 0).order_by(attendance.c.time)
        ).fetchall()

    if not unsynced:
        return jsonify({"ok": True, "sent": 0, "message": "All records already synced"})

    lines = []
    ids   = []
    for r in unsynced:
        d = dict(r._mapping)
        t = d.get("time")
        if not t:
            continue
        sid  = str(d["staff_id"]).strip().upper()
        sint = STATUS_TO_INT.get(d.get("status", "Check In"), 0)
        lines.append(f"ATTLOG\t{sid}\t{t.strftime('%Y-%m-%d %H:%M:%S')}\t{sint}\t1\t0\t0")
        ids.append(d["id"])

    if not lines:
        return jsonify({"ok": True, "sent": 0, "message": "No valid records"})

    body = urllib.parse.urlencode({"data": "\n".join(lines)}).encode("utf-8")
    url  = f"{RAILWAY_URL}/iclock/cdata?SN={DEVICE_SERIAL}&table=ATTLOG"
    try:
        req = urllib.request.Request(url, data=body,
            headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            if resp.status != 200:
                return jsonify({"ok": False, "error": f"Railway {resp.status}"}), 502
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

    with engine.connect() as conn:
        conn.execute(update(attendance).where(attendance.c.id.in_(ids)).values(synced=1))
        conn.commit()

    return jsonify({"ok": True, "sent": len(lines), "message": f"{len(lines)} record(s) uploaded to Railway"})

@app.route("/api/sync/status")
def api_sync_status():
    with engine.connect() as conn:
        total    = conn.execute(select(attendance)).fetchall()
        unsynced = conn.execute(select(attendance).where(attendance.c.synced == 0)).fetchall()
    return jsonify({"total": len(total), "synced": len(total) - len(unsynced), "pending": len(unsynced)})

# ── ADMS Protocol ──────────────────────────────────────────────────────────────

@app.route("/iclock/cdata", methods=["GET"])
def adms_register():
    sn  = request.args.get("SN", "UNKNOWN")
    fw  = request.args.get("firmware", "")
    mark_device_seen(sn, fw, ip=request.remote_addr)
    print(f"\n[DEVICE REGISTERED]  SN={sn}  IP={request.remote_addr}")
    socketio.emit("device_status", {"serial": sn, "online": True, "ip": request.remote_addr,
                                     "last_seen": device["last_seen"].isoformat()})
    return (
        f"GET OPTION FROM: {sn}\n"
        "ATTLOGStamp=9999\nOPERLOGStamp=9999\nATTPHOTOStamp=9999\n"
        "ErrorDelay=30\nDelay=10\nTransTimes=00:00;23:59\nTransInterval=1\n"
        "TransFlag=0011011000000000\nRealtime=1\nEncrypt=None\n"
        "ServerVer=2.4.1 2015-04-14\nPushProtVer=2.4.1 2015-04-14\nPushOptionsFlag=1\n"
    )

@app.route("/iclock/cdata", methods=["POST"])
def adms_receive():
    sn  = request.args.get("SN", device["serial"] or "UNKNOWN")
    tbl = request.args.get("table", "")
    mark_device_seen(sn, ip=request.remote_addr)

    raw = request.get_data(as_text=True).strip()
    print(f"\n[ADMS POST]  SN={sn}  table={tbl}\n{raw}")

    if tbl and tbl.upper() not in ("", "ATTLOG"):
        return "OK"

    for line in raw.splitlines():
        line = line.strip()
        if not line or line.startswith("OPLOG"):
            continue

        if line.startswith("ATTLOG"):
            # Format A: ATTLOG\tPIN\tYYYY-MM-DD HH:MM:SS\tSTATUS\tVERIFY\t...
            parts   = line.split("\t")
            if len(parts) < 3: continue
            sid     = str(parts[1]).strip().upper()
            try:    ts = datetime.datetime.strptime(parts[2].strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError: continue
            ptype   = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 0
        else:
            # Format B (MB160 raw): PIN  DATE  TIME  VERIFY  STATUS  ...
            parts  = line.split()
            if len(parts) < 3: continue
            sid    = str(parts[0]).strip().upper()
            try:   ts = datetime.datetime.strptime(parts[1] + " " + parts[2], "%Y-%m-%d %H:%M:%S")
            except ValueError: continue
            ptype  = int(parts[4]) if len(parts) > 4 and parts[4].isdigit() else 0

        status_label = STATUS_MAP.get(ptype, "Check In")

        with engine.connect() as conn:
            emp = conn.execute(select(employees).where(employees.c.staff_id == sid)).fetchone()
            if emp:
                conn.execute(attendance.insert().values(
                    staff_id=sid, name=emp.name, time=ts,
                    status=status_label, synced=0
                ))
                conn.commit()
                payload = {"staff_id": sid, "name": emp.name, "department": emp.department,
                           "position": emp.position, "time": ts.strftime("%H:%M:%S"),
                           "date": ts.strftime("%Y-%m-%d"), "status": status_label}
                print(f"  ✅  {emp.name}  ({sid})  {ts.strftime('%H:%M:%S')}  {status_label}")
                socketio.emit("attendance", payload)
            else:
                print(f"  ❌  Unknown: {sid}")
                socketio.emit("unknown_scan", {"staff_id": sid, "time": ts.strftime("%H:%M:%S")})

    return "OK"

@app.route("/iclock/getrequest", methods=["GET"])
def adms_heartbeat():
    mark_device_seen(request.args.get("SN") or None, ip=request.remote_addr)
    return "OK"

@app.route("/iclock/devicecmd", methods=["POST"])
def adms_devicecmd():
    return "OK"

# ── Standalone start ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    cfg  = _load_cfg()
    port = cfg.get("port", 80)
    try:
        lan_ip = _socket.gethostbyname(_socket.gethostname())
    except Exception:
        lan_ip = "127.0.0.1"
    print(f"\n{'='*52}\n  ZKTeco Local Attendance Server\n{'='*52}")
    print(f"  Dashboard  →  http://localhost:{port}")
    print(f"  Device IP  →  set device Server Address: {lan_ip}")
    print(f"  Cloud URL  →  {RAILWAY_URL}")
    print(f"{'='*52}\n")
    socketio.run(app, host="0.0.0.0", port=port, debug=False, allow_unsafe_werkzeug=True)
